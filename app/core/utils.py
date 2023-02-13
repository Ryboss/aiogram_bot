import openpyxl
import csv
import io

from typing import Any, List, Set, TypeVar, Type, Union, Callable, Generator, Iterable, Tuple, Optional, Literal
from aiogram import types

from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.worksheet.dimensions import DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.numbers import FORMAT_NUMBER, FORMAT_NUMBER_00, FORMAT_PERCENTAGE_00
from datetime import datetime, timedelta

from app.models.utils import GenericValidationError, ExcelModel

MSK = timedelta(hours=3)


def sql_validation_error(loc: List[str], type: str, msg: Optional[str] = None) -> List[dict]:
    """
    Генерирует сообщение с ошибками SQL возвращаемое в стиле ошибок pydantic'а, когда ошибка идёт на уровне БД, например:
    UniqueViolationError, ForeignKeyViolationError по каким-то полям
    """

    return [GenericValidationError(loc=loc, msg=msg, type=type).dict(exclude_none=True)]


async def generate_excel(file_path: str, message: types.Message) -> str:
    """
    Генерация Excel по файлу CSV
    """

    wb = openpyxl.Workbook()
    ws = wb.active
    dh: DimensionHolder = ws.column_dimensions

    border = Side(border_style=BORDER_THIN, color="00000000")

    thin_border = Border(
        top=border,
        left=border,
        right=border,
        bottom=border,)

    with open(file_path) as f:
        reader = csv.reader(f, delimiter=",")
        for i, row in enumerate(reader):
            ws.append(row)
            for s, column in enumerate(row):
                ws.cell(row=i + 1, column=s + 1).border = thin_border
                dh[get_column_letter(s + 1)].width = len(column) + 5

    path_to_save = f"download_files/{message.from_user.id}/generate_excel_{message.message_id}.xlsx"
    wb.save(filename=path_to_save)

    return path_to_save


async def generate_json_to_excel(data: List[ExcelModel], file_path: str) -> str:
    """
    Генерация Excel на основе json
    """

    with io.BytesIO() as stream:
        wb = openpyxl.Workbook()

        border = Side(border_style=BORDER_THIN, color="00000000")
        data_for_excel = []

        r_len = len(data)

        new_list = []

        thin_border = Border(
            top=border,
            left=border,
            right=border,
            bottom=border,)

        for i, sheet in enumerate(data):
            title = sheet.title
            titles = sheet.titles
            data_new = sheet.data

            lcol = len(titles)
            lrow = len(data_new)

            ws: Worksheet = wb.create_sheet(f"Users{i + 1}") if r_len > 1 and i > 0 else wb.active
            ws.title = title
            dh: DimensionHolder = ws.column_dimensions  # type: ignore
            for i, t in enumerate(titles):
                if t:
                    dh[get_column_letter(i + 1)].width = len(t) + 5
                    ws.cell(column=i + 1, row=1, value=t).border = thin_border

            for info in data_new:
                for key in info.dict().values():
                    new_list.append(key)

                data_for_excel.append(new_list)
                new_list = []

            values = zip(ws.iter_rows(min_row=2, max_row=lrow + 2, min_col=1, max_col=lcol), data_for_excel)

            for rows, data_row in values:
                for cell, val in zip(rows, data_row):
                    if isinstance(val, int):
                        cell.value = val
                        cell.number_format = FORMAT_NUMBER
                    elif isinstance(val, float):
                        cell.value = val
                        cell.number_format = FORMAT_NUMBER_00
                    elif isinstance(val, str) and val.endswith("%"):
                        cell.value = float(val.rstrip("%"))
                        cell.number_format = FORMAT_PERCENTAGE_00
                    elif isinstance(val, datetime):
                        cell.value = val + MSK  # Excel does not have timezone
                    elif val is not None:
                        cell.value = val


                    cell.border = thin_border

        wb.save(filename=file_path)
        stream.seek(0)
        return file_path
